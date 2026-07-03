from __future__ import annotations

import csv
import hashlib
import json
import re
from dataclasses import dataclass, field
from datetime import UTC, datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from sqlalchemy import select
from sqlalchemy.orm import Session

from src.modules.supplier_registry.models import SupplierContact, SupplierExternalRef, SupplierProfile, SupplierTag
from src.shared.db.base import utcnow
from src.shared.enums import SupplierStatus
from src.shared.ids import next_supplier_id
from src.shared.validation import require_non_empty


HEADER_ALIASES: dict[str, list[str]] = {
    "legal_name": [
        "legal_name",
        "Юрлицо",
        "Юридическое лицо",
        "Наименование",
        "Название компании",
        "Поставщик",
    ],
    "display_name": [
        "display_name",
        "Краткое название",
        "Название",
        "Отображаемое название",
    ],
    "inn": ["inn", "ИНН"],
    "website": ["website", "url", "Сайт"],
    "email": ["email", "Email", "Почта", "E-mail"],
    "phone": ["phone", "Телефон"],
    "categories": ["categories", "Категории", "Категория", "Направления", "Товары"],
    "brands": ["brands", "Бренды", "Производители", "Марки"],
    "region": ["region", "Регион", "Город", "География"],
    "notes": ["notes", "Комментарий", "Примечание", "Заметки"],
}

TAG_TRANSLITERATION = str.maketrans(
    {
        "а": "a",
        "б": "b",
        "в": "v",
        "г": "g",
        "д": "d",
        "е": "e",
        "ё": "e",
        "ж": "zh",
        "з": "z",
        "и": "i",
        "й": "y",
        "к": "k",
        "л": "l",
        "м": "m",
        "н": "n",
        "о": "o",
        "п": "p",
        "р": "r",
        "с": "s",
        "т": "t",
        "у": "u",
        "ф": "f",
        "х": "h",
        "ц": "ts",
        "ч": "ch",
        "ш": "sh",
        "щ": "sch",
        "ъ": "",
        "ы": "y",
        "ь": "",
        "э": "e",
        "ю": "yu",
        "я": "ya",
    }
)


@dataclass(slots=True)
class VendorImportIssue:
    row_number: int
    message: str


@dataclass(slots=True)
class VendorImportSummary:
    source_label: str
    source_file: str
    total_rows: int = 0
    created_suppliers: int = 0
    updated_suppliers: int = 0
    skipped_rows: int = 0
    rows_without_inn: int = 0
    possible_duplicates: int = 0
    contacts_created: int = 0
    tags_created: int = 0
    errors: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[dict[str, Any]] = field(default_factory=list)
    imported_at: str = field(default_factory=lambda: datetime.now(UTC).isoformat())

    def to_dict(self) -> dict[str, Any]:
        return {
            "source_label": self.source_label,
            "source_file": self.source_file,
            "total_rows": self.total_rows,
            "created_suppliers": self.created_suppliers,
            "updated_suppliers": self.updated_suppliers,
            "skipped_rows": self.skipped_rows,
            "rows_without_inn": self.rows_without_inn,
            "possible_duplicates": self.possible_duplicates,
            "contacts_created": self.contacts_created,
            "tags_created": self.tags_created,
            "errors": self.errors,
            "warnings": self.warnings,
            "imported_at": self.imported_at,
        }


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[^a-zа-я0-9]+", "", text)


HEADER_LOOKUP: dict[str, str] = {
    _normalize_header(alias): canonical
    for canonical, aliases in HEADER_ALIASES.items()
    for alias in aliases
}


def _transliterate(text: str) -> str:
    lowered = text.lower()
    output: list[str] = []
    for char in lowered:
        mapped = TAG_TRANSLITERATION.get(ord(char))
        if mapped is not None:
            output.append(mapped)
        else:
            output.append(char)
    return "".join(output)


def _canonical_tag(prefix: str, value: str) -> str:
    ascii_like = _transliterate(value)
    normalized = re.sub(r"[^a-zA-Z0-9]+", "_", ascii_like).strip("_").upper()
    normalized = re.sub(r"_+", "_", normalized)
    if not normalized:
        normalized = "UNSPECIFIED"
    return f"{prefix}_{normalized}"


def _split_multi_value(value: str | None) -> list[str]:
    if not value:
        return []
    parts = re.split(r"[,;\n\r]+", value)
    return [item.strip() for item in parts if item and item.strip()]


def _normalize_inn(value: str | None) -> str:
    return re.sub(r"\D+", "", value or "")


def _normalize_email(value: str | None) -> str | None:
    email = (value or "").strip().lower()
    return email or None


def _normalize_phone(value: str | None) -> str | None:
    raw = (value or "").strip()
    if not raw:
        return None
    digits = re.sub(r"\D+", "", raw)
    if len(digits) == 11 and digits.startswith("7"):
        return f"+7 {digits[1:4]} {digits[4:7]}-{digits[7:9]}-{digits[9:11]}"
    if len(digits) == 11 and digits.startswith("8"):
        return f"+7 {digits[1:4]} {digits[4:7]}-{digits[7:9]}-{digits[9:11]}"
    if len(digits) == 10:
        return f"+7 {digits[0:3]} {digits[3:6]}-{digits[6:8]}-{digits[8:10]}"
    return raw


def _normalize_website(value: str | None) -> tuple[str | None, str | None]:
    raw = (value or "").strip()
    if not raw:
        return None, None
    candidate = raw if "://" in raw else f"https://{raw}"
    parsed = urlparse(candidate)
    host = parsed.netloc.strip().lower()
    if not host:
        return None, None
    host = host.lstrip("www.")
    normalized = f"{parsed.scheme or 'https'}://{host}{parsed.path or ''}".rstrip("/")
    return normalized, host


def _email_domain(email: str | None) -> str | None:
    if not email or "@" not in email:
        return None
    return email.split("@", 1)[1].strip().lower()


def _display_or_legal_name(row: dict[str, Any]) -> tuple[str | None, str | None]:
    legal_name = (row.get("legal_name") or "").strip()
    display_name = (row.get("display_name") or "").strip()
    fallback_name = (
        legal_name
        or display_name
        or str(row.get("name") or "").strip()
        or str(row.get("поставщик") or "").strip()
    )
    if not legal_name:
        legal_name = fallback_name
    if not display_name:
        display_name = legal_name
    return legal_name or None, display_name or None


def _stable_missing_inn(row: dict[str, Any], source_label: str, row_number: int) -> str:
    fingerprint = "|".join(
        [
            source_label,
            str(row_number),
            str(row.get("legal_name") or ""),
            str(row.get("display_name") or ""),
            str(row.get("website") or ""),
            str(row.get("email") or ""),
        ]
    )
    digest = hashlib.sha1(fingerprint.encode("utf-8")).hexdigest()[:12].upper()
    return f"NOINN-{digest}"


def _read_csv_rows(file_path: Path) -> list[dict[str, Any]]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            with file_path.open("r", encoding=encoding, newline="") as handle:
                return list(csv.DictReader(handle))
        except UnicodeDecodeError as exc:
            last_error = exc
    raise ValueError(f"Could not decode CSV file '{file_path}': {last_error}")


def _read_xlsx_rows(file_path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(item or "").strip() for item in rows[0]]
    result: list[dict[str, Any]] = []
    for row in rows[1:]:
        result.append({headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))})
    return result


def _canonicalize_row(raw_row: dict[str, Any]) -> dict[str, Any]:
    canonical: dict[str, Any] = {}
    for header, value in raw_row.items():
        mapped = HEADER_LOOKUP.get(_normalize_header(header))
        if mapped:
            canonical[mapped] = value
    return canonical


def _load_vendor_rows(file_path: Path) -> list[dict[str, Any]]:
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        rows = _read_csv_rows(file_path)
    elif suffix == ".xlsx":
        rows = _read_xlsx_rows(file_path)
    else:
        raise ValueError(f"Unsupported file format: {file_path.suffix}")
    return [_canonicalize_row(row) for row in rows]


def _normalize_name_for_similarity(value: str | None) -> str:
    return re.sub(r"[^a-zа-я0-9]+", " ", _transliterate((value or "").lower())).strip()


def _supplier_domains(session: Session, supplier_id: str) -> set[str]:
    domains: set[str] = set()
    refs = session.scalars(select(SupplierExternalRef).where(SupplierExternalRef.supplier_id == supplier_id)).all()
    for ref in refs:
        if ref.ref_type == "website":
            _normalized, domain = _normalize_website(ref.ref_value)
            if domain:
                domains.add(domain)
    contacts = session.scalars(select(SupplierContact).where(SupplierContact.supplier_id == supplier_id)).all()
    for contact in contacts:
        domain = _email_domain(contact.email)
        if domain:
            domains.add(domain)
    return domains


def _possible_duplicate_matches(
    session: Session,
    *,
    supplier: SupplierProfile | None,
    legal_name: str,
    display_name: str,
    website_domain: str | None,
    email_domain: str | None,
) -> list[dict[str, Any]]:
    domains = {item for item in {website_domain, email_domain} if item}
    if not domains:
        return []
    normalized_target = _normalize_name_for_similarity(display_name or legal_name)
    results: list[dict[str, Any]] = []
    candidates = list(session.scalars(select(SupplierProfile).order_by(SupplierProfile.created_at.asc())))
    for candidate in candidates:
        if supplier is not None and candidate.supplier_id == supplier.supplier_id:
            continue
        candidate_domains = _supplier_domains(session, candidate.supplier_id)
        if not domains.intersection(candidate_domains):
            continue
        similarity = max(
            SequenceMatcher(None, normalized_target, _normalize_name_for_similarity(candidate.display_name)).ratio(),
            SequenceMatcher(None, normalized_target, _normalize_name_for_similarity(candidate.legal_name)).ratio(),
        )
        if similarity >= 0.82:
            results.append(
                {
                    "supplier_id": candidate.supplier_id,
                    "supplier_name": candidate.display_name,
                    "matched_domains": sorted(domains.intersection(candidate_domains)),
                    "name_similarity": round(similarity, 3),
                }
            )
    return results


def _ensure_external_ref(session: Session, supplier_id: str, ref_type: str, ref_value: str) -> bool:
    existing = session.scalar(
        select(SupplierExternalRef).where(
            SupplierExternalRef.supplier_id == supplier_id,
            SupplierExternalRef.ref_type == ref_type,
            SupplierExternalRef.ref_value == ref_value,
        )
    )
    if existing:
        return False
    session.add(SupplierExternalRef(supplier_id=supplier_id, ref_type=ref_type, ref_value=ref_value))
    return True


def _ensure_contact(
    session: Session,
    supplier_id: str,
    *,
    contact_name: str,
    email: str | None,
    phone: str | None,
) -> bool:
    if not email and not phone:
        return False
    contacts = list(session.scalars(select(SupplierContact).where(SupplierContact.supplier_id == supplier_id)))
    for existing in contacts:
        if (email and existing.email == email) or (phone and existing.phone == phone):
            return False
    is_primary = not any(contact.is_primary for contact in contacts)
    session.add(
        SupplierContact(
            supplier_id=supplier_id,
            contact_name=contact_name,
            email=email,
            phone=phone,
            is_primary=is_primary,
        )
    )
    return True


def _ensure_tag(session: Session, supplier_id: str, tag_code: str) -> bool:
    existing = session.scalar(
        select(SupplierTag).where(SupplierTag.supplier_id == supplier_id, SupplierTag.tag_code == tag_code)
    )
    if existing:
        return False
    session.add(SupplierTag(supplier_id=supplier_id, tag_code=tag_code))
    return True


def _merge_notes(existing_notes: str | None, new_notes: str | None) -> str | None:
    parts = [item.strip() for item in [existing_notes or "", new_notes or ""] if item and item.strip()]
    if not parts:
        return None
    deduped: list[str] = []
    for item in parts:
        if item not in deduped:
            deduped.append(item)
    return "\n".join(deduped)


def import_vendor_list(
    session: Session,
    *,
    operator_id: str,
    file_path: Path,
    source_label: str,
) -> VendorImportSummary:
    if not file_path.is_file():
        raise FileNotFoundError(f"Vendor list file was not found: {file_path}")

    rows = _load_vendor_rows(file_path)
    summary = VendorImportSummary(source_label=source_label, source_file=str(file_path))
    summary.total_rows = len(rows)
    operator_tag = _canonical_tag("OPERATOR", operator_id)

    for row_index, raw_row in enumerate(rows, start=2):
        try:
            legal_name, display_name = _display_or_legal_name(raw_row)
            if not legal_name or not display_name:
                summary.skipped_rows += 1
                summary.errors.append({"row_number": row_index, "message": "Row skipped because supplier name is missing"})
                continue

            inn = _normalize_inn(str(raw_row.get("inn") or ""))
            email = _normalize_email(str(raw_row.get("email") or ""))
            phone = _normalize_phone(str(raw_row.get("phone") or ""))
            website_url, website_domain = _normalize_website(str(raw_row.get("website") or ""))
            email_domain = _email_domain(email)
            notes = str(raw_row.get("notes") or "").strip() or None
            categories = _split_multi_value(str(raw_row.get("categories") or ""))
            brands = _split_multi_value(str(raw_row.get("brands") or ""))
            regions = _split_multi_value(str(raw_row.get("region") or ""))

            synthetic_inn = False
            if not inn:
                synthetic_inn = True
                inn = _stable_missing_inn(raw_row, source_label, row_index)
                summary.rows_without_inn += 1
                summary.warnings.append({"row_number": row_index, "message": "INN is missing; supplier tagged for review"})

            supplier = session.scalar(select(SupplierProfile).where(SupplierProfile.inn == inn))
            created = supplier is None
            if supplier is None:
                supplier = SupplierProfile(
                    supplier_id=next_supplier_id(session, SupplierProfile.supplier_id),
                    legal_name=require_non_empty(legal_name, "legal_name"),
                    display_name=require_non_empty(display_name, "display_name"),
                    inn=require_non_empty(inn, "inn"),
                    country_code="RU",
                    status=SupplierStatus.DRAFT if synthetic_inn else SupplierStatus.ACTIVE,
                    notes=notes,
                )
                session.add(supplier)
                session.flush()
                summary.created_suppliers += 1
            else:
                supplier.legal_name = legal_name or supplier.legal_name
                supplier.display_name = display_name or supplier.display_name
                if synthetic_inn and supplier.status == SupplierStatus.ACTIVE:
                    supplier.status = SupplierStatus.DRAFT
                supplier.notes = _merge_notes(supplier.notes, notes)
                supplier.updated_at = utcnow()
                session.add(supplier)
                summary.updated_suppliers += 1

            if website_url:
                _ensure_external_ref(session, supplier.supplier_id, "website", website_url)
            _ensure_external_ref(session, supplier.supplier_id, "vendor_list_source", f"{source_label}|{file_path.name}")
            _ensure_external_ref(session, supplier.supplier_id, "vendor_list_row", str(row_index))

            if _ensure_contact(session, supplier.supplier_id, contact_name="Sales", email=email, phone=phone):
                summary.contacts_created += 1

            tag_codes = {
                "SOURCE_VENDOR_LIST",
                operator_tag,
            }
            if not created:
                tag_codes.add("SOURCE_VENDOR_LIST_ENRICHED")
            if synthetic_inn:
                tag_codes.add("NEEDS_REVIEW_NO_INN")
            for category in categories:
                tag_codes.add(_canonical_tag("CATEGORY", category))
            for brand in brands:
                tag_codes.add(_canonical_tag("BRAND", brand))
            for region in regions:
                tag_codes.add(_canonical_tag("REGION", region))

            duplicates = _possible_duplicate_matches(
                session,
                supplier=supplier,
                legal_name=legal_name,
                display_name=display_name,
                website_domain=website_domain,
                email_domain=email_domain,
            )
            if duplicates:
                summary.possible_duplicates += len(duplicates)
                tag_codes.add("POSSIBLE_DUPLICATE_VENDOR_LIST")
                summary.warnings.append(
                    {
                        "row_number": row_index,
                        "message": "Possible duplicate candidates found",
                        "matches": duplicates,
                    }
                )

            for tag_code in sorted(tag_codes):
                if _ensure_tag(session, supplier.supplier_id, tag_code):
                    summary.tags_created += 1

            session.commit()
        except Exception as exc:
            session.rollback()
            summary.skipped_rows += 1
            summary.errors.append({"row_number": row_index, "message": str(exc)})

    return summary


def build_vendor_import_report_markdown(summary: VendorImportSummary) -> str:
    lines = [
        "# Vendor Import Report",
        "",
        f"- Source label: {summary.source_label}",
        f"- Source file: {summary.source_file}",
        f"- Imported at: {summary.imported_at}",
        f"- Total rows: {summary.total_rows}",
        f"- Created suppliers: {summary.created_suppliers}",
        f"- Updated suppliers: {summary.updated_suppliers}",
        f"- Skipped rows: {summary.skipped_rows}",
        f"- Rows without INN: {summary.rows_without_inn}",
        f"- Possible duplicates: {summary.possible_duplicates}",
        f"- Contacts created: {summary.contacts_created}",
        f"- Tags created: {summary.tags_created}",
        "",
        "## Warnings",
        "",
    ]
    if summary.warnings:
        for item in summary.warnings:
            suffix = ""
            if item.get("matches"):
                suffix = f" | matches={json.dumps(item['matches'], ensure_ascii=False)}"
            lines.append(f"- row {item['row_number']}: {item['message']}{suffix}")
    else:
        lines.append("- None")

    lines += ["", "## Errors", ""]
    if summary.errors:
        for item in summary.errors:
            lines.append(f"- row {item['row_number']}: {item['message']}")
    else:
        lines.append("- None")
    return "\n".join(lines)
