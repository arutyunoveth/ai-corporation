from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import UTC, datetime
from difflib import SequenceMatcher
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from sqlalchemy import select
from sqlalchemy.orm import Session

from src.modules.quote_repository.tkp_normalization_schemas import (
    NormalizedTKPLineItem,
    NormalizedTKPQuote,
    NormalizedTKPQuoteBatch,
)
from src.modules.supplier_registry.models import SupplierContact, SupplierExternalRef, SupplierProfile
from src.shared.enums import SupplierStatus


HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "item_name": ("Наименование", "Товар", "Позиция", "Описание", "Номенклатура", "item", "name", "description"),
    "manufacturer": ("Производитель", "manufacturer"),
    "brand": ("Бренд", "brand", "Марка"),
    "model": ("Модель", "model"),
    "article": ("Артикул", "article", "sku"),
    "quantity": ("Кол-во", "Количество", "qty", "quantity"),
    "unit": ("Ед.", "Ед. изм.", "Единица", "unit"),
    "unit_price": ("Цена", "Цена за ед.", "Цена без НДС", "unit_price", "price"),
    "total_price": ("Сумма", "Итого", "Стоимость", "total", "amount"),
    "vat": ("НДС", "VAT", "Ставка НДС"),
    "delivery_time_days": ("Срок поставки", "Срок", "delivery", "delivery_time"),
    "warranty": ("Гарантия", "warranty"),
    "payment_terms": ("Условия оплаты", "Оплата", "payment_terms"),
    "availability": ("Наличие", "availability", "stock"),
    "supplier_label": ("Поставщик", "Supplier", "supplier"),
    "inn": ("ИНН", "inn"),
    "email": ("Email", "E-mail", "Почта", "email"),
    "phone": ("Телефон", "phone"),
}

_RUS_TRANSLIT = str.maketrans(
    {
        "а": "a",
        "б": "b",
        "в": "v",
        "г": "g",
        "д": "d",
        "е": "e",
        "ё": "e",
        "ж": "zh",
        "з": "z",
        "и": "i",
        "й": "y",
        "к": "k",
        "л": "l",
        "м": "m",
        "н": "n",
        "о": "o",
        "п": "p",
        "р": "r",
        "с": "s",
        "т": "t",
        "у": "u",
        "ф": "f",
        "х": "h",
        "ц": "ts",
        "ч": "ch",
        "ш": "sh",
        "щ": "sch",
        "ъ": "",
        "ы": "y",
        "ь": "",
        "э": "e",
        "ю": "yu",
        "я": "ya",
    }
)

BOOL_TRUE_HINTS = ("yes", "true", "included", "включ", "есть", "available", "in stock", "да", "подтверж")
BOOL_FALSE_HINTS = ("no", "false", "not available", "не включ", "нет", "pending", "made to order")
TEXT_EXTENSIONS = {".txt", ".md"}
SPREADSHEET_EXTENSIONS = {".xlsx", ".csv"}


@dataclass(slots=True)
class _DeterministicParseResult:
    quote: NormalizedTKPQuote
    llm_hint_text: str
    inn: str | None
    email_domain: str | None
    website_domain: str | None
    supplier_name_for_match: str
    fuzzy_supplier_name: str | None = None


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-zа-яё0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


HEADER_LOOKUP = {
    _normalize_header(alias): canonical
    for canonical, aliases in HEADER_ALIASES.items()
    for alias in aliases
}


def _normalize_name(value: str) -> str:
    translated = value.lower().translate(_RUS_TRANSLIT)
    translated = re.sub(r"[^a-z0-9]+", " ", translated)
    return re.sub(r"\s+", " ", translated).strip()


def _normalize_supplier_label(file_path: Path) -> str:
    stem = file_path.stem
    stem = re.sub(r"(?i)\b(tkp|кп|tkp|quote|offer|supplier|поставщик)\b", " ", stem)
    stem = re.sub(r"[_\-]+", " ", stem)
    cleaned = re.sub(r"\s+", " ", stem).strip()
    return cleaned or file_path.stem


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    normalized = (
        text.replace("\u00a0", "")
        .replace(" ", "")
        .replace("₽", "")
        .replace("руб.", "")
        .replace("руб", "")
        .replace("RUB", "")
        .replace("USD", "")
        .replace("EUR", "")
    )
    cleaned = re.sub(r"[^0-9,.\-]", "", normalized)
    if not cleaned:
        return None
    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif cleaned.count(",") > 1 and "." not in cleaned:
        cleaned = cleaned.replace(",", "")
    elif cleaned.count(".") > 1 and "," not in cleaned:
        cleaned = cleaned.replace(".", "")
    elif cleaned.count(",") == 1 and len(cleaned.split(",")[-1]) == 3 and len(cleaned) > 4:
        cleaned = cleaned.replace(",", "")
    elif cleaned.count(".") == 1 and len(cleaned.split(".")[-1]) == 3 and len(cleaned) > 4:
        cleaned = cleaned.replace(".", "")
    else:
        cleaned = cleaned.replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


def _extract_days(value: Any) -> int | None:
    text = str(value or "").strip().lower()
    if not text:
        return None
    match = re.search(r"(\d{1,4})", text)
    return int(match.group(1)) if match else None


def _extract_months(value: Any) -> int | None:
    text = str(value or "").strip().lower()
    if not text:
        return None
    match = re.search(r"(\d{1,3})", text)
    return int(match.group(1)) if match else None


def _extract_date(value: str | None) -> str | None:
    text = str(value or "").strip()
    if not text:
        return None
    for pattern in (r"(\d{4}-\d{2}-\d{2})", r"(\d{2}\.\d{2}\.\d{4})"):
        match = re.search(pattern, text)
        if match:
            return match.group(1)
    return None


def _extract_currency(text: str) -> str | None:
    upper = text.upper()
    if "RUB" in upper or "РУБ" in upper or "₽" in text:
        return "RUB"
    if "USD" in upper or "$" in text:
        return "USD"
    if "EUR" in upper or "€" in text:
        return "EUR"
    return None


def _extract_bool(value: str | None) -> bool | None:
    normalized = str(value or "").strip().lower()
    if not normalized:
        return None
    if any(hint in normalized for hint in BOOL_TRUE_HINTS):
        return True
    if any(hint in normalized for hint in BOOL_FALSE_HINTS):
        return False
    return None


def _extract_vat_rate(value: str | None) -> float | None:
    text = str(value or "").strip().lower()
    if not text:
        return None
    match = re.search(r"(\d{1,2}(?:[.,]\d+)?)\s*%", text)
    if match:
        return float(match.group(1).replace(",", "."))
    if "без ндс" in text or "without vat" in text:
        return 0.0
    return None


def _extract_email(text: str) -> str | None:
    match = re.search(r"([A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,})", text, re.IGNORECASE)
    return match.group(1).lower() if match else None


def _extract_phone(text: str) -> str | None:
    match = re.search(r"(\+?\d[\d() \-]{8,}\d)", text)
    return match.group(1).strip() if match else None


def _email_domain(email: str | None) -> str | None:
    if not email or "@" not in email:
        return None
    return email.split("@", 1)[1].strip().lower()


def _extract_urls(text: str) -> list[str]:
    return re.findall(r"https?://[^\s)]+", text, flags=re.IGNORECASE)


def _website_domain(url: str | None) -> str | None:
    if not url:
        return None
    candidate = url if "://" in url else f"https://{url}"
    parsed = urlparse(candidate)
    host = parsed.netloc.lower().lstrip("www.")
    return host or None


def _extract_inn(text: str) -> str | None:
    match = re.search(r"(?:инн|inn)\D{0,6}(\d{10,12})", text, re.IGNORECASE)
    if match:
        return match.group(1)
    return None


def _parse_text_quote_fields(text: str) -> dict[str, Any]:
    compact = text.replace("\r", "")
    total_amount = _to_float(_find_field(compact, ("total price", "итого", "total amount", "сумма", "стоимость")))
    total_without_vat = _to_float(_find_field(compact, ("price without vat", "без ндс", "total without vat")))
    unit_price = _to_float(_find_field(compact, ("price per unit", "цена за ед", "unit price", "цена")))
    return {
        "quote_date": _extract_date(_find_field(compact, ("quote date", "дата", "date"))),
        "valid_until": _extract_date(_find_field(compact, ("valid until", "offer validity", "действительно до"))),
        "currency_code": _extract_currency(compact) or "RUB",
        "vat_rate": _extract_vat_rate(_find_field(compact, ("vat", "ндс"))),
        "vat_included": _extract_bool(_find_field(compact, ("with vat", "с ндс", "включая ндс"))),
        "total_amount": total_amount,
        "total_amount_without_vat": total_without_vat,
        "delivery_cost": _to_float(_find_field(compact, ("delivery cost", "стоимость доставки", "доставка"))),
        "delivery_time_days": _extract_days(_find_field(compact, ("delivery time", "срок поставки", "срок"))),
        "payment_terms": _find_field(compact, ("payment terms", "условия оплаты", "оплата")),
        "warranty_months": _extract_months(_find_field(compact, ("warranty", "гарантия"))),
        "availability": _find_field(compact, ("availability", "наличие")),
        "includes_delivery": _extract_bool(_find_field(compact, ("packaging/delivery", "delivery included", "доставка включ"))),
        "includes_installation": _extract_bool(_find_field(compact, ("installation", "монтаж"))),
        "certificates_available": _extract_bool(_find_field(compact, ("certificate", "сертификат", "eac", "iso"))),
        "contact_email": _extract_email(compact),
        "contact_phone": _extract_phone(compact),
        "unit_price": unit_price,
    }


def _find_field(text: str, labels: tuple[str, ...]) -> str | None:
    for label in labels:
        pattern = rf"{re.escape(label)}\s*[:\-]\s*(.+)"
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1).splitlines()[0].strip()
    return None


def _read_csv_rows(file_path: Path) -> list[dict[str, Any]]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            with file_path.open("r", encoding=encoding, newline="") as handle:
                return list(csv.DictReader(handle))
        except UnicodeDecodeError as exc:
            last_error = exc
    raise ValueError(f"Could not decode CSV file '{file_path}': {last_error}")


def _read_xlsx_rows(file_path: Path) -> list[list[Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    rows: list[list[Any]] = []
    for sheet in workbook.worksheets:
        rows.extend([list(row) for row in sheet.iter_rows(values_only=True)])
    return rows


def _tabular_preview_from_rows(headers: list[str], rows: list[dict[str, Any]], *, max_rows: int = 20) -> str:
    preview_lines = [" | ".join(headers)]
    for row in rows[:max_rows]:
        preview_lines.append(" | ".join(str(row.get(header, "") or "") for header in headers))
    return "\n".join(preview_lines)


def build_tkp_llm_inputs(tkp_files: list[Path]) -> list[dict[str, str]]:
    inputs: list[dict[str, str]] = []
    for file_path in tkp_files:
        suffix = file_path.suffix.lower()
        supplier_label = _normalize_supplier_label(file_path)
        if suffix in TEXT_EXTENSIONS:
            quote_text = file_path.read_text(encoding="utf-8")
        elif suffix == ".csv":
            rows = _read_csv_rows(file_path)
            headers = list(rows[0].keys()) if rows else []
            quote_text = _tabular_preview_from_rows(headers, rows) if headers else ""
        elif suffix == ".xlsx":
            raw_rows = _read_xlsx_rows(file_path)
            if not raw_rows:
                quote_text = ""
            else:
                headers = [str(value or "") for value in raw_rows[0]]
                normalized_rows = [
                    {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}
                    for row in raw_rows[1:21]
                ]
                quote_text = _tabular_preview_from_rows(headers, normalized_rows) if headers else ""
        else:
            continue
        inputs.append(
            {
                "supplier_label": supplier_label,
                "source_file": str(file_path),
                "quote_text": quote_text[:8000],
            }
        )
    return inputs


def _canonicalize_row(raw_row: dict[str, Any]) -> dict[str, Any]:
    canonical: dict[str, Any] = {}
    for header, value in raw_row.items():
        mapped = HEADER_LOOKUP.get(_normalize_header(header))
        if mapped:
            canonical[mapped] = value
    return canonical


def _score_spreadsheet_header(row: list[Any]) -> tuple[int, dict[str, int]]:
    mapping: dict[str, int] = {}
    for index, value in enumerate(row):
        mapped = HEADER_LOOKUP.get(_normalize_header(value))
        if mapped and mapped not in mapping:
            mapping[mapped] = index
    score = len(mapping)
    if "item_name" in mapping:
        score += 2
    if "quantity" in mapping:
        score += 1
    if "unit_price" in mapping or "total_price" in mapping:
        score += 2
    return score, mapping


def _find_best_header_row(rows: list[list[Any]]) -> tuple[int | None, dict[str, int]]:
    best_score = 0
    best_index: int | None = None
    best_mapping: dict[str, int] = {}
    for idx, row in enumerate(rows[:20]):
        score, mapping = _score_spreadsheet_header(row)
        if score > best_score:
            best_score = score
            best_index = idx
            best_mapping = mapping
    return best_index, best_mapping


def _build_review_fields(quote: NormalizedTKPQuote) -> list[str]:
    fields: list[str] = []
    if quote.total_amount is None:
        fields.append("total_amount")
    if quote.delivery_time_days is None:
        fields.append("delivery_time_days")
    if quote.payment_terms is None:
        fields.append("payment_terms")
    if not quote.line_items:
        fields.append("line_items")
    for item in quote.line_items:
        if item.quantity is None:
            fields.append(f"line_items[{item.item_index}].quantity")
        if item.unit_price is None and item.total_price is None:
            fields.append(f"line_items[{item.item_index}].pricing")
    return sorted(set(fields))


def _base_quote(file_path: Path, supplier_label: str, parser_mode: str) -> NormalizedTKPQuote:
    return NormalizedTKPQuote(
        supplier_label=supplier_label,
        source_file=str(file_path),
        normalization_status="needs_review",
        extraction_confidence=0.0,
        parser_mode=parser_mode,  # type: ignore[arg-type]
        human_review_required=True,
    )


def _parse_text_file(file_path: Path) -> _DeterministicParseResult:
    text = file_path.read_text(encoding="utf-8")
    supplier_label = _normalize_supplier_label(file_path)
    fields = _parse_text_quote_fields(text)
    quote = _base_quote(file_path, supplier_label, "deterministic")
    quote.quote_date = fields["quote_date"]
    quote.valid_until = fields["valid_until"]
    quote.currency_code = fields["currency_code"]
    quote.vat_included = fields["vat_included"]
    quote.vat_rate = fields["vat_rate"]
    quote.total_amount = fields["total_amount"]
    quote.total_amount_without_vat = fields["total_amount_without_vat"]
    quote.delivery_cost = fields["delivery_cost"]
    quote.delivery_time_days = fields["delivery_time_days"]
    quote.payment_terms = fields["payment_terms"]
    quote.warranty_months = fields["warranty_months"]
    quote.availability = fields["availability"]
    quote.includes_delivery = fields["includes_delivery"]
    quote.includes_installation = fields["includes_installation"]
    quote.certificates_available = fields["certificates_available"]
    quote.contact_email = fields["contact_email"]
    quote.contact_phone = fields["contact_phone"]
    quote.line_items = [
        NormalizedTKPLineItem(
            item_index=1,
            item_name="Commercial offer summary",
            unit_price=fields["unit_price"],
            total_price=fields["total_amount"],
            delivery_time_days=fields["delivery_time_days"],
            vat_included=fields["vat_included"],
            compliance_status="needs_review",
            comments="Derived from text summary.",
            confidence=0.62,
        )
    ] if any(
        [
            fields["unit_price"] is not None,
            fields["total_amount"] is not None,
            fields["delivery_time_days"] is not None,
        ]
    ) else []
    confidence = 0.28
    for key in ("total_amount", "delivery_time_days", "payment_terms", "warranty_months", "availability"):
        if fields[key] is not None:
            confidence += 0.12
    if quote.contact_email:
        confidence += 0.06
    if quote.line_items:
        confidence += 0.06
    if quote.currency_code == "RUB" and not _extract_currency(text):
        quote.warnings.append("currency_assumed_rub")
    if quote.vat_included is None and quote.vat_rate is None:
        quote.warnings.append("vat_unknown")
    quote.extraction_confidence = min(round(confidence, 3), 0.92)
    quote.fields_needing_review = _build_review_fields(quote)
    quote.normalization_status = "parsed" if quote.extraction_confidence >= 0.7 else "needs_review"
    if not quote.total_amount and not quote.line_items:
        quote.normalization_status = "failed"
        quote.fields_needing_review.append("line_items")
    inn = _extract_inn(text)
    urls = _extract_urls(text)
    return _DeterministicParseResult(
        quote=quote,
        llm_hint_text=text[:8000],
        inn=inn,
        email_domain=_email_domain(quote.contact_email),
        website_domain=_website_domain(urls[0]) if urls else None,
        supplier_name_for_match=quote.supplier_label,
    )


def _parse_csv_or_xlsx(file_path: Path) -> _DeterministicParseResult:
    supplier_label = _normalize_supplier_label(file_path)
    quote = _base_quote(file_path, supplier_label, "deterministic")

    if file_path.suffix.lower() == ".csv":
        raw_rows = _read_csv_rows(file_path)
        rows = [list(raw_rows[0].keys())] + [list(row.values()) for row in raw_rows] if raw_rows else []
        structured_rows = [_canonicalize_row(row) for row in raw_rows]
    else:
        rows = _read_xlsx_rows(file_path)
        header_idx, header_mapping = _find_best_header_row(rows)
        if header_idx is None:
            header_idx = 0
            header_mapping = {}
        headers = [str(value or "") for value in rows[header_idx]] if rows else []
        raw_rows = [
            {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}
            for row in rows[header_idx + 1 :]
        ] if headers else []
        structured_rows = [_canonicalize_row(row) for row in raw_rows]

    line_items: list[NormalizedTKPLineItem] = []
    supplier_name = supplier_label
    quote_level_text_parts = [supplier_label]
    inn: str | None = None

    for index, row in enumerate(structured_rows, start=1):
        if not any(value not in (None, "", " ") for value in row.values()):
            continue
        if row.get("supplier_label") and not supplier_name:
            supplier_name = str(row["supplier_label"]).strip()
        if row.get("inn") and not inn:
            matched = re.sub(r"\D+", "", str(row["inn"]))
            inn = matched or inn
        if row.get("email") and not quote.contact_email:
            quote.contact_email = str(row["email"]).strip().lower()
        if row.get("phone") and not quote.contact_phone:
            quote.contact_phone = str(row["phone"]).strip()
        item_name = str(row.get("item_name") or "").strip()
        quantity = _to_float(row.get("quantity"))
        unit_price = _to_float(row.get("unit_price"))
        total_price = _to_float(row.get("total_price"))
        if total_price is None and quantity is not None and unit_price is not None:
            total_price = round(quantity * unit_price, 2)
        if not item_name and not any(value is not None for value in (quantity, unit_price, total_price)):
            quote_level_text_parts.append(" | ".join(f"{k}: {v}" for k, v in row.items() if v not in (None, "")))
            continue

        delivery_days = _extract_days(row.get("delivery_time_days"))
        item = NormalizedTKPLineItem(
            item_index=len(line_items) + 1,
            item_name=item_name or f"Line item {index}",
            manufacturer=str(row.get("manufacturer")).strip() if row.get("manufacturer") else None,
            brand=str(row.get("brand")).strip() if row.get("brand") else None,
            model=str(row.get("model")).strip() if row.get("model") else None,
            article=str(row.get("article")).strip() if row.get("article") else None,
            quantity=quantity,
            unit=str(row.get("unit")).strip() if row.get("unit") else None,
            unit_price=unit_price,
            total_price=total_price,
            vat_included=_extract_bool(str(row.get("vat") or "")),
            delivery_time_days=delivery_days,
            compliance_status="needs_review",
            comments=None,
            confidence=min(
                0.32
                + (0.2 if item_name else 0.0)
                + (0.14 if quantity is not None else 0.0)
                + (0.18 if unit_price is not None or total_price is not None else 0.0)
                + (0.08 if delivery_days is not None else 0.0),
                0.94,
            ),
        )
        line_items.append(item)
        quote_level_text_parts.append(" | ".join(f"{k}: {v}" for k, v in row.items() if v not in (None, "")))
        if quote.delivery_time_days is None:
            quote.delivery_time_days = delivery_days
        if quote.payment_terms is None and row.get("payment_terms"):
            quote.payment_terms = str(row.get("payment_terms")).strip()
        if quote.warranty_months is None and row.get("warranty"):
            quote.warranty_months = _extract_months(row.get("warranty"))
        if quote.availability is None and row.get("availability"):
            quote.availability = str(row.get("availability")).strip()

    quote.supplier_label = supplier_name or supplier_label
    quote.line_items = line_items
    quote.total_amount = round(sum(item.total_price for item in line_items if item.total_price is not None), 2) if line_items else None
    if quote.currency_code == "RUB":
        quote.warnings.append("currency_assumed_rub")
    quote.warnings.append("vat_unknown")
    parsed_text = "\n".join(quote_level_text_parts)
    text_fields = _parse_text_quote_fields(parsed_text)
    quote.currency_code = text_fields["currency_code"] or "RUB"
    quote.vat_included = text_fields["vat_included"]
    quote.vat_rate = text_fields["vat_rate"]
    quote.delivery_cost = text_fields["delivery_cost"]
    quote.total_amount_without_vat = text_fields["total_amount_without_vat"]
    quote.payment_terms = quote.payment_terms or text_fields["payment_terms"]
    quote.warranty_months = quote.warranty_months or text_fields["warranty_months"]
    quote.availability = quote.availability or text_fields["availability"]
    quote.includes_delivery = text_fields["includes_delivery"]
    quote.includes_installation = text_fields["includes_installation"]
    quote.certificates_available = text_fields["certificates_available"]
    quote.contact_email = quote.contact_email or text_fields["contact_email"]
    quote.contact_phone = quote.contact_phone or text_fields["contact_phone"]
    if quote.total_amount is None:
        quote.total_amount = text_fields["total_amount"]
    if quote.delivery_time_days is None:
        quote.delivery_time_days = text_fields["delivery_time_days"]
    quote.extraction_confidence = round(
        min(
            0.24
            + (0.22 if line_items else 0.0)
            + (0.18 if quote.total_amount is not None else 0.0)
            + (0.12 if quote.delivery_time_days is not None else 0.0)
            + (0.12 if quote.payment_terms else 0.0)
            + (0.08 if quote.contact_email else 0.0),
            0.95,
        ),
        3,
    )
    quote.fields_needing_review = _build_review_fields(quote)
    quote.normalization_status = "parsed" if line_items and quote.extraction_confidence >= 0.72 else "needs_review"
    if not line_items:
        quote.normalization_status = "failed"
    return _DeterministicParseResult(
        quote=quote,
        llm_hint_text=parsed_text[:8000],
        inn=inn,
        email_domain=_email_domain(quote.contact_email),
        website_domain=_website_domain(_extract_urls(parsed_text)[0]) if _extract_urls(parsed_text) else None,
        supplier_name_for_match=quote.supplier_label,
    )


def _unsupported_quote(file_path: Path) -> _DeterministicParseResult:
    supplier_label = _normalize_supplier_label(file_path)
    quote = NormalizedTKPQuote(
        supplier_label=supplier_label,
        supplier_id=None,
        source_file=str(file_path),
        normalization_status="unsupported_format" if file_path.suffix.lower() == ".pdf" else "failed",
        quote_date=None,
        valid_until=None,
        currency_code="RUB",
        vat_included=None,
        vat_rate=None,
        total_amount=None,
        total_amount_without_vat=None,
        delivery_cost=None,
        delivery_time_days=None,
        payment_terms=None,
        warranty_months=None,
        availability=None,
        includes_delivery=None,
        includes_installation=None,
        certificates_available=None,
        contact_email=None,
        contact_phone=None,
        line_items=[],
        extraction_confidence=0.0,
        fields_needing_review=["source_file", "line_items", "supplier_id"],
        warnings=["requires_manual_text_extract" if file_path.suffix.lower() == ".pdf" else "unsupported_format"],
        parser_mode="deterministic",
        human_review_required=True,
    )
    return _DeterministicParseResult(
        quote=quote,
        llm_hint_text="",
        inn=None,
        email_domain=None,
        website_domain=None,
        supplier_name_for_match=supplier_label,
    )


def parse_tkp_file(file_path: Path) -> _DeterministicParseResult:
    suffix = file_path.suffix.lower()
    if suffix in TEXT_EXTENSIONS:
        return _parse_text_file(file_path)
    if suffix in SPREADSHEET_EXTENSIONS:
        return _parse_csv_or_xlsx(file_path)
    return _unsupported_quote(file_path)


def _validated_llm_quotes(llm_quotes: list[dict[str, Any]] | None) -> list[NormalizedTKPQuote]:
    if not llm_quotes:
        return []
    batch = NormalizedTKPQuoteBatch.model_validate({"quotes": llm_quotes})
    return batch.quotes


def _quote_lookup_key(*, source_file: str | None, supplier_label: str | None) -> tuple[str, str]:
    source_key = str(source_file or "").strip().lower()
    label_key = _normalize_name(str(supplier_label or ""))
    return source_key, label_key


def _merge_quotes(deterministic: NormalizedTKPQuote, llm_quote: NormalizedTKPQuote | None) -> NormalizedTKPQuote:
    if llm_quote is None:
        return deterministic
    if deterministic.normalization_status == "parsed" and deterministic.extraction_confidence >= 0.95:
        return deterministic

    merged = deterministic.model_copy(deep=True)
    for field_name in (
        "quote_date",
        "valid_until",
        "currency_code",
        "vat_included",
        "vat_rate",
        "total_amount",
        "total_amount_without_vat",
        "delivery_cost",
        "delivery_time_days",
        "payment_terms",
        "warranty_months",
        "availability",
        "includes_delivery",
        "includes_installation",
        "certificates_available",
        "contact_email",
        "contact_phone",
    ):
        if getattr(merged, field_name) in (None, "", []):
            setattr(merged, field_name, getattr(llm_quote, field_name))
    if not merged.line_items and llm_quote.line_items:
        merged.line_items = llm_quote.line_items
    merged.fields_needing_review = sorted(set(merged.fields_needing_review + llm_quote.fields_needing_review))
    merged.warnings = list(dict.fromkeys(merged.warnings + llm_quote.warnings))
    merged.normalization_status = llm_quote.normalization_status if llm_quote.normalization_status != "failed" else merged.normalization_status
    merged.extraction_confidence = max(merged.extraction_confidence, llm_quote.extraction_confidence)
    merged.parser_mode = "hybrid" if merged.parser_mode == "deterministic" else llm_quote.parser_mode
    return merged


def _supplier_contact_domains(session: Session, supplier_id: str) -> set[str]:
    domains: set[str] = set()
    contacts = session.scalars(select(SupplierContact).where(SupplierContact.supplier_id == supplier_id)).all()
    for contact in contacts:
        domain = _email_domain(contact.email)
        if domain:
            domains.add(domain)
    refs = session.scalars(select(SupplierExternalRef).where(SupplierExternalRef.supplier_id == supplier_id)).all()
    for ref in refs:
        if ref.ref_type == "website":
            domain = _website_domain(ref.ref_value)
            if domain:
                domains.add(domain)
    return domains


def _match_supplier(
    session: Session,
    *,
    quote: NormalizedTKPQuote,
    inn: str | None,
    email_domain: str | None,
    website_domain: str | None,
    supplier_name: str,
) -> None:
    if inn:
        supplier = session.scalar(select(SupplierProfile).where(SupplierProfile.inn == inn))
        if supplier:
            quote.supplier_id = supplier.supplier_id
            return

    candidates = list(
        session.scalars(
            select(SupplierProfile)
            .where(SupplierProfile.status.in_([SupplierStatus.ACTIVE, SupplierStatus.DRAFT]))
            .order_by(SupplierProfile.created_at.asc(), SupplierProfile.id.asc())
        )
    )
    normalized_name = _normalize_name(supplier_name)
    exact_name = [
        supplier
        for supplier in candidates
        if normalized_name and normalized_name in {_normalize_name(supplier.display_name), _normalize_name(supplier.legal_name)}
    ]
    if len(exact_name) == 1:
        quote.supplier_id = exact_name[0].supplier_id
        return

    domain = email_domain or website_domain
    if domain:
        matched_by_domain = [
            supplier for supplier in candidates if domain in _supplier_contact_domains(session, supplier.supplier_id)
        ]
        if len(matched_by_domain) == 1:
            quote.supplier_id = matched_by_domain[0].supplier_id
            return

    best_match: SupplierProfile | None = None
    best_score = 0.0
    for supplier in candidates:
        similarity = max(
            SequenceMatcher(None, normalized_name, _normalize_name(supplier.display_name)).ratio(),
            SequenceMatcher(None, normalized_name, _normalize_name(supplier.legal_name)).ratio(),
        )
        if similarity > best_score:
            best_match = supplier
            best_score = similarity
    if best_match and best_score >= 0.9:
        quote.warnings.append(f"possible_supplier_match:{best_match.supplier_id}")
    quote.fields_needing_review.append("supplier_id")


def normalize_tkp_quotes(
    session: Session,
    *,
    tkp_files: list[Path],
    llm_quotes: list[dict[str, Any]] | None = None,
) -> list[NormalizedTKPQuote]:
    llm_validated = _validated_llm_quotes(llm_quotes)
    llm_lookup = {
        _quote_lookup_key(source_file=quote.source_file, supplier_label=quote.supplier_label): quote
        for quote in llm_validated
    }
    normalized_quotes: list[NormalizedTKPQuote] = []
    for file_path in tkp_files:
        parsed = parse_tkp_file(file_path)
        merged = _merge_quotes(
            parsed.quote,
            llm_lookup.get(
                _quote_lookup_key(source_file=str(file_path), supplier_label=parsed.quote.supplier_label)
            )
            or llm_lookup.get(_quote_lookup_key(source_file=None, supplier_label=parsed.quote.supplier_label)),
        )
        merged.fields_needing_review = sorted(set(_build_review_fields(merged) + merged.fields_needing_review))
        _match_supplier(
            session,
            quote=merged,
            inn=parsed.inn,
            email_domain=parsed.email_domain,
            website_domain=parsed.website_domain,
            supplier_name=parsed.supplier_name_for_match,
        )
        if merged.supplier_id is None and "supplier_id" not in merged.fields_needing_review:
            merged.fields_needing_review.append("supplier_id")
        normalized_quotes.append(merged)
    return normalized_quotes


def build_tkp_normalization_report(quotes: list[NormalizedTKPQuote]) -> str:
    lines = [
        "# TKP Normalization Report",
        "",
        f"- Generated at: {datetime.now(UTC).isoformat()}",
        f"- Files processed: {len(quotes)}",
        "",
    ]
    for quote in quotes:
        lines += [
            f"## {Path(quote.source_file).name}",
            "",
            f"- Supplier label: {quote.supplier_label}",
            f"- Supplier ID: {quote.supplier_id or 'not matched'}",
            f"- Status: {quote.normalization_status}",
            f"- Parser mode: {quote.parser_mode}",
            f"- Confidence: {quote.extraction_confidence}",
            f"- Total amount: {quote.total_amount if quote.total_amount is not None else 'unknown'} {quote.currency_code}",
            f"- Delivery time: {quote.delivery_time_days if quote.delivery_time_days is not None else 'unknown'}",
            f"- Payment terms: {quote.payment_terms or 'unknown'}",
            f"- Line items: {len(quote.line_items)}",
            f"- Fields needing review: {', '.join(quote.fields_needing_review) if quote.fields_needing_review else 'none'}",
            f"- Warnings: {', '.join(quote.warnings) if quote.warnings else 'none'}",
            "",
        ]
    return "\n".join(lines)


def build_tkp_comparison_from_normalized_quotes(
    quotes: list[NormalizedTKPQuote],
    *,
    analysis_mode: str,
    method: str,
) -> dict[str, Any]:
    suppliers: list[dict[str, Any]] = []
    for quote in quotes:
        suppliers.append(
            {
                "supplier_label": quote.supplier_label,
                "supplier_id": quote.supplier_id,
                "source_file": quote.source_file,
                "status": quote.normalization_status,
                "parser_mode": quote.parser_mode,
                "currency": quote.currency_code,
                "total_amount": quote.total_amount,
                "total_amount_without_vat": quote.total_amount_without_vat,
                "delivery_cost": quote.delivery_cost,
                "delivery_time_days": quote.delivery_time_days,
                "warranty_months": quote.warranty_months,
                "payment_terms": quote.payment_terms,
                "availability": quote.availability,
                "includes_delivery": quote.includes_delivery,
                "includes_installation": quote.includes_installation,
                "certificates_available": quote.certificates_available,
                "contact_email": quote.contact_email,
                "contact_phone": quote.contact_phone,
                "line_items_count": len(quote.line_items),
                "fields_needing_review": quote.fields_needing_review,
                "warnings": quote.warnings,
                "extraction_confidence": quote.extraction_confidence,
                "human_review_required": quote.human_review_required,
            }
        )

    parsed_quotes = [quote for quote in quotes if quote.normalization_status == "parsed"]
    needs_review_quotes = [quote for quote in quotes if quote.normalization_status == "needs_review"]
    return {
        "suppliers": suppliers,
        "comparison_generated_at": datetime.now(UTC).isoformat(),
        "method": method,
        "analysis_mode": analysis_mode,
        "normalized_quotes_count": len(quotes),
        "parsed_quotes_count": len(parsed_quotes),
        "needs_review_quotes_count": len(needs_review_quotes),
        "supplier_quotes_found": len([quote for quote in quotes if quote.total_amount is not None]),
        "items_extracted": sum(len(quote.line_items) for quote in quotes),
        "status": "needs_review" if needs_review_quotes else "completed",
        "note": "Normalized from supplier TKP files. Human review remains required.",
    }
