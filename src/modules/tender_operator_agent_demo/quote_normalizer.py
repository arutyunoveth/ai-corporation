from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from statistics import median
from typing import Any

from src.modules.tender_operator_agent_demo.schemas import (
    EconomicsSummary,
    ExtractionWarning,
    ManualCheck,
    QuoteComparison,
    QuoteItem,
    QuoteOffer,
    SupplierQuote,
)


HEADER_SYNONYMS: dict[str, tuple[str, ...]] = {
    "item_number": ("№", "n", "номер", "п/п", "item", "row"),
    "name": (
        "наименование",
        "товар",
        "позиция",
        "описание",
        "номенклатура",
        "оборудование",
        "description",
        "product",
        "name",
        "item",
    ),
    "quantity": ("кол-во", "количество", "qty", "quantity"),
    "unit": ("ед", "ед изм", "ед. изм", "единица", "unit", "uom"),
    "unit_price": ("цена", "цена за ед", "unit price", "price", "price per unit", "цена без ндс", "цена с ндс"),
    "total_price": ("сумма", "стоимость", "итого", "total", "amount", "sum"),
    "delivery": ("срок", "срок поставки", "поставка", "delivery", "delivery days", "delivery date"),
    "manufacturer": ("производитель", "бренд", "марка", "manufacturer", "brand"),
    "currency": ("валюта", "currency"),
}

SUPPLIER_QUOTE_HINTS = ("кп", "ткп", "quote", "offer", "поставщик", "supplier", "коммерческое предложение")
CUSTOMER_SPEC_HINTS = ("спецификация", "тз", "ведомость", "заказчик", "specification", "statement")


@dataclass
class SpreadsheetSource:
    file_id: str
    display_name: str
    source_file: str
    extension: str
    raw_content: bytes
    source: str


@dataclass
class ParsedSpreadsheet:
    source: SpreadsheetSource
    document_type: str
    supplier_name: str | None
    items: list[QuoteItem]
    warnings: list[str]
    limitations: list[str]
    currency: str | None
    delivery_summary: str | None


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("\n", " ").replace("\t", " ")
    text = re.sub(r"[^a-zа-яё0-9%]+", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _normalize_name(value: str) -> str:
    text = _normalize_header(value)
    text = re.sub(r"\b(шт|pcs|pc|компл|set|набор|ед)\b", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _match_header(header: str) -> str | None:
    if not header:
        return None
    for canonical, synonyms in HEADER_SYNONYMS.items():
        for synonym in synonyms:
            normalized_synonym = _normalize_header(synonym)
            if not normalized_synonym:
                continue
            if header == normalized_synonym:
                return canonical
            if len(normalized_synonym) >= 3 and normalized_synonym in header:
                return canonical
    return None


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("\u00a0", "").replace(" ", "")
    text = text.replace("₽", "").replace("руб.", "").replace("руб", "")
    text = text.replace("usd", "").replace("eur", "").replace("$", "").replace("€", "")
    text = text.replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def _to_currency(value: Any) -> str | None:
    text = str(value or "").strip().upper()
    if not text:
        return None
    if "RUB" in text or "РУБ" in text or "₽" in text:
        return "RUB"
    if "USD" in text or "$" in text:
        return "USD"
    if "EUR" in text or "€" in text:
        return "EUR"
    return None


def _detect_supplier_name(file_name: str, sheet_rows: list[list[Any]]) -> str:
    top_text = " ".join(str(cell or "") for row in sheet_rows[:8] for cell in row[:6])
    match = re.search(r"(?:supplier|поставщик)\s*[:\-]?\s*([A-Za-zА-Яа-яЁё0-9 _\"'().-]{3,80})", top_text, re.IGNORECASE)
    if match:
        return match.group(1).strip().strip(":;-")
    stem = Path(file_name).stem
    cleaned = re.sub(r"(?i)\b(tkp|kp|quote|offer|supplier|поставщик|коммерческое|предложение)\b", " ", stem)
    cleaned = re.sub(r"[_\-]+", " ", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned or stem


def _detect_document_type(file_name: str, rows: list[list[Any]], headers: dict[str, int]) -> str:
    lowered_name = file_name.lower()
    visible_text = " ".join(str(cell or "") for row in rows[:12] for cell in row[:8]).lower()
    has_price = any(key in headers for key in ("unit_price", "total_price")) or "цена" in visible_text or "стоимость" in visible_text
    has_qty = "quantity" in headers or "количество" in visible_text or "qty" in visible_text

    if has_price and (has_qty or "name" in headers):
        if any(token in lowered_name for token in SUPPLIER_QUOTE_HINTS) or "коммерческое предложение" in visible_text:
            return "supplier_quote"
        return "price_table"
    if has_qty and not has_price and (any(token in lowered_name for token in CUSTOMER_SPEC_HINTS) or "спецификация" in visible_text):
        return "customer_specification"
    return "unknown_table"


def _score_header_row(values: list[Any]) -> tuple[int, dict[str, int]]:
    mapping: dict[str, int] = {}
    for index, value in enumerate(values):
        matched = _match_header(_normalize_header(value))
        if matched and matched not in mapping:
            mapping[matched] = index
    score = len(mapping)
    if "name" in mapping:
        score += 2
    if "quantity" in mapping:
        score += 1
    if "unit_price" in mapping or "total_price" in mapping:
        score += 2
    return score, mapping


def _find_header_row(rows: list[list[Any]]) -> tuple[int | None, dict[str, int]]:
    best_score = 0
    best_row: int | None = None
    best_mapping: dict[str, int] = {}
    for idx, row in enumerate(rows[:20]):
        score, mapping = _score_header_row(row)
        if score > best_score:
            best_score = score
            best_row = idx
            best_mapping = mapping
    return best_row, best_mapping


def _safe_openpyxl_workbook(content: bytes):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError(f"openpyxl is unavailable: {exc}") from exc
    return load_workbook(BytesIO(content), data_only=True, read_only=True)


def parse_spreadsheet_source(source: SpreadsheetSource) -> ParsedSpreadsheet:
    if source.extension == ".xls":
        return ParsedSpreadsheet(
            source=source,
            document_type="unknown_table",
            supplier_name=None,
            items=[],
            warnings=["Legacy .xls parsing is limited in this demo sprint."],
            limitations=["Use .xlsx when possible for deterministic quote extraction."],
            currency=None,
            delivery_summary=None,
        )

    workbook = _safe_openpyxl_workbook(source.raw_content)
    all_items: list[QuoteItem] = []
    warnings: list[str] = []
    limitations: list[str] = []
    currency_values: list[str] = []
    delivery_values: list[str] = []
    collected_rows: list[list[Any]] = []

    for sheet in workbook.worksheets:
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        if not rows:
            continue
        collected_rows.extend(rows[:8])
        header_row_idx, mapping = _find_header_row(rows)
        if header_row_idx is None or "name" not in mapping:
            warnings.append(f"Sheet '{sheet.title}' did not produce a reliable header row.")
            continue

        blank_streak = 0
        for row_index in range(header_row_idx + 1, min(len(rows), header_row_idx + 250)):
            row = rows[row_index]
            name_value = row[mapping["name"]] if mapping["name"] < len(row) else None
            quantity = _to_float(row[mapping["quantity"]]) if "quantity" in mapping and mapping["quantity"] < len(row) else None
            unit_price = _to_float(row[mapping["unit_price"]]) if "unit_price" in mapping and mapping["unit_price"] < len(row) else None
            total_price = _to_float(row[mapping["total_price"]]) if "total_price" in mapping and mapping["total_price"] < len(row) else None
            if not any([name_value, quantity, unit_price, total_price]):
                blank_streak += 1
                if blank_streak >= 4:
                    break
                continue
            blank_streak = 0

            normalized_name = _normalize_name(str(name_value or ""))
            if len(normalized_name) < 2 and not (unit_price or total_price):
                continue

            unit = str(row[mapping["unit"]]).strip() if "unit" in mapping and mapping["unit"] < len(row) and row[mapping["unit"]] else None
            manufacturer = (
                str(row[mapping["manufacturer"]]).strip()
                if "manufacturer" in mapping and mapping["manufacturer"] < len(row) and row[mapping["manufacturer"]]
                else None
            )
            delivery = (
                str(row[mapping["delivery"]]).strip()
                if "delivery" in mapping and mapping["delivery"] < len(row) and row[mapping["delivery"]]
                else None
            )
            currency = (
                _to_currency(row[mapping["currency"]])
                if "currency" in mapping and mapping["currency"] < len(row)
                else None
            )
            if not total_price and unit_price is not None and quantity is not None:
                total_price = unit_price * quantity
            if total_price is not None and unit_price is None and quantity not in (None, 0):
                unit_price = total_price / quantity

            row_warnings: list[str] = []
            if quantity is None:
                row_warnings.append("Quantity was not confidently extracted.")
            if unit_price is None and total_price is None:
                row_warnings.append("Price was not confidently extracted.")
            if unit is None:
                row_warnings.append("Unit of measure was not found.")

            confidence = 0.3
            if normalized_name:
                confidence += 0.25
            if quantity is not None:
                confidence += 0.15
            if unit_price is not None or total_price is not None:
                confidence += 0.2
            if unit:
                confidence += 0.05
            if manufacturer:
                confidence += 0.05
            if delivery:
                confidence += 0.05
            confidence = min(confidence, 0.98)

            item = QuoteItem(
                item_number=str(row[mapping["item_number"]]).strip() if "item_number" in mapping and mapping["item_number"] < len(row) and row[mapping["item_number"]] else None,
                row_number=row_index + 1,
                normalized_name=normalized_name or f"row-{row_index + 1}",
                requested_quantity=quantity,
                unit=unit,
                manufacturer=manufacturer,
                currency=currency,
                source_file=source.source_file,
                source_sheet=sheet.title,
                confidence=confidence,
                warnings=row_warnings,
            )
            item.offers.append(
                QuoteOffer(
                    supplier_name=_detect_supplier_name(source.display_name, rows),
                    offered_name=str(name_value or "").strip() or item.normalized_name,
                    quantity=quantity,
                    unit=unit,
                    unit_price=unit_price,
                    total_price=total_price,
                    delivery=delivery,
                    confidence=confidence,
                    warnings=row_warnings,
                )
            )
            all_items.append(item)
            if currency:
                currency_values.append(currency)
            if delivery:
                delivery_values.append(delivery)

    document_type = _detect_document_type(source.display_name, collected_rows, _find_header_row(collected_rows)[1] if collected_rows else {})
    supplier_name = _detect_supplier_name(source.display_name, collected_rows) if all_items else None
    currency = currency_values[0] if currency_values else None
    delivery_summary = ", ".join(sorted(set(delivery_values))[:3]) if delivery_values else None
    if not all_items:
        warnings.append(f"No structured quote rows were extracted from '{source.display_name}'.")
        limitations.append("Spreadsheet layout is unsupported or too irregular for this deterministic parser.")

    return ParsedSpreadsheet(
        source=source,
        document_type=document_type,
        supplier_name=supplier_name,
        items=all_items,
        warnings=list(dict.fromkeys(warnings)),
        limitations=list(dict.fromkeys(limitations)),
        currency=currency,
        delivery_summary=delivery_summary,
    )


def _build_supplier_quote(parsed: ParsedSpreadsheet, supplier_index: int) -> SupplierQuote:
    totals = [offer.total_price for item in parsed.items for offer in item.offers if offer.total_price is not None]
    price_conf = median([item.confidence for item in parsed.items]) if parsed.items else 0.0
    completeness = 0.0
    if parsed.items:
        complete_rows = 0
        for item in parsed.items:
            offer = item.offers[0]
            if item.normalized_name and (offer.unit_price is not None or offer.total_price is not None):
                complete_rows += 1
        completeness = complete_rows / len(parsed.items)
    return SupplierQuote(
        supplier_id=f"SUP-{supplier_index:02d}",
        supplier_name=parsed.supplier_name or f"Supplier {supplier_index}",
        source_file=parsed.source.source_file,
        source_sheet=parsed.items[0].source_sheet if parsed.items else None,
        document_type=parsed.document_type,
        total_amount=sum(totals) if totals else None,
        currency=parsed.currency,
        items_count=len(parsed.items),
        delivery_summary=parsed.delivery_summary,
        completeness_score=round(completeness, 3),
        price_confidence=round(price_conf, 3),
        warnings=parsed.warnings,
        items=parsed.items,
    )


def build_quote_comparison(spreadsheet_sources: list[SpreadsheetSource], analysis_mode: str) -> QuoteComparison:
    if not spreadsheet_sources:
        return QuoteComparison(
            status="blocked",
            analysis_mode=analysis_mode,
            supplier_quotes_found=0,
            items_extracted=0,
            suppliers=[],
            items=[],
            comparison_summary={},
            manual_checks=[ManualCheck(code="quotes_missing", message="ТКП не загружены. Сравнение поставщиков недоступно.")],
            warnings=[],
            limitations=["Supplier quote files were not provided."],
        )

    parsed_results: list[ParsedSpreadsheet] = []
    top_warnings: list[ExtractionWarning] = []
    limitations: list[str] = []
    for source in spreadsheet_sources:
        try:
            parsed = parse_spreadsheet_source(source)
        except Exception as exc:
            top_warnings.append(ExtractionWarning(code="spreadsheet_parse_failed", message=f"{source.display_name}: {exc}"))
            limitations.append(f"Spreadsheet '{source.display_name}' could not be parsed deterministically.")
            continue
        parsed_results.append(parsed)
        top_warnings.extend(
            ExtractionWarning(code="spreadsheet_warning", message=f"{source.display_name}: {message}")
            for message in parsed.warnings
        )
        limitations.extend(parsed.limitations)

    supplier_quotes = [
        _build_supplier_quote(parsed, index)
        for index, parsed in enumerate(parsed_results, start=1)
        if parsed.document_type in {"supplier_quote", "price_table"}
    ]
    spec_items = [
        item
        for parsed in parsed_results
        if parsed.document_type == "customer_specification"
        for item in parsed.items
    ]

    if not supplier_quotes:
        return QuoteComparison(
            status="needs_review" if parsed_results else "blocked",
            analysis_mode=analysis_mode,
            supplier_quotes_found=0,
            items_extracted=0,
            suppliers=[],
            items=[],
            comparison_summary={},
            manual_checks=[ManualCheck(code="quotes_not_recognized", message="Excel files were loaded, but supplier quote tables were not confidently recognized.")],
            warnings=top_warnings,
            limitations=list(dict.fromkeys(limitations + ["Supplier quote classification confidence was insufficient."])),
        )

    requested_quantities = {item.normalized_name: item.requested_quantity for item in spec_items if item.normalized_name}
    grouped: dict[str, QuoteItem] = {}
    for supplier in supplier_quotes:
        for item in supplier.items:
            key = item.normalized_name
            existing = grouped.get(key)
            if existing is None:
                grouped[key] = QuoteItem(
                    item_number=item.item_number,
                    row_number=item.row_number,
                    normalized_name=item.normalized_name,
                    requested_quantity=requested_quantities.get(key, item.requested_quantity),
                    unit=item.unit,
                    manufacturer=item.manufacturer,
                    currency=item.currency or supplier.currency,
                    source_file=item.source_file,
                    source_sheet=item.source_sheet,
                    confidence=item.confidence,
                    warnings=list(item.warnings),
                    offers=[],
                    needs_review=False,
                )
                existing = grouped[key]
            existing.offers.extend(item.offers)
            if item.unit and existing.unit and item.unit != existing.unit:
                existing.needs_review = True
                existing.warnings.append("Units differ across offers.")

    for item in grouped.values():
        prices = [(offer.supplier_name, offer.total_price or offer.unit_price) for offer in item.offers if (offer.total_price or offer.unit_price) is not None]
        if prices:
            prices.sort(key=lambda pair: pair[1] or float("inf"))
            item.best_price_supplier = prices[0][0]
            if len(prices) > 1 and prices[0][1]:
                item.price_spread_percent = round(((prices[-1][1] - prices[0][1]) / prices[0][1]) * 100, 2)
        if any(offer.confidence < 0.55 for offer in item.offers):
            item.needs_review = True
        if len({offer.unit for offer in item.offers if offer.unit}) > 1:
            item.needs_review = True

    comparison_items = sorted(grouped.values(), key=lambda item: item.normalized_name)
    mixed_currencies = {supplier.currency for supplier in supplier_quotes if supplier.currency}
    manual_checks: list[ManualCheck] = []
    if any(item.needs_review for item in comparison_items):
        manual_checks.append(ManualCheck(code="low_confidence_items", message="Есть позиции с низкой уверенностью или несовпадающими единицами измерения."))
    if len(mixed_currencies) > 1:
        manual_checks.append(ManualCheck(code="mixed_currencies", message="В ТКП обнаружены разные валюты. Требуется ручная нормализация."))
    if any(not supplier.delivery_summary for supplier in supplier_quotes):
        manual_checks.append(ManualCheck(code="delivery_missing", message="Не у всех поставщиков найден срок поставки."))

    summary = {
        "best_total_supplier": min(
            (
                {"supplier_name": supplier.supplier_name, "total_amount": supplier.total_amount}
                for supplier in supplier_quotes
                if supplier.total_amount is not None
            ),
            default=None,
            key=lambda item: item["total_amount"],
        ),
        "suppliers_with_prices": len([supplier for supplier in supplier_quotes if supplier.total_amount is not None]),
        "items_with_best_price": len([item for item in comparison_items if item.best_price_supplier]),
    }
    status = "completed"
    if manual_checks or top_warnings:
        status = "partial"
    if not comparison_items:
        status = "needs_review"

    return QuoteComparison(
        status=status,
        analysis_mode=analysis_mode,
        supplier_quotes_found=len(supplier_quotes),
        items_extracted=len(comparison_items),
        suppliers=supplier_quotes,
        items=comparison_items,
        comparison_summary=summary,
        manual_checks=manual_checks,
        warnings=top_warnings,
        limitations=list(dict.fromkeys(limitations)),
    )


def build_economics_summary(
    *,
    quote_comparison: QuoteComparison,
    analysis_mode: str,
    target_margin_percent: float,
    logistics_reserve_percent: float,
    risk_reserve_percent: float,
    payment_delay_days: int,
) -> EconomicsSummary:
    if not quote_comparison.suppliers:
        return EconomicsSummary(
            status="blocked",
            analysis_mode=analysis_mode,
            currency=None,
            supplier_cost_min=None,
            supplier_cost_selected=None,
            expected_revenue=None,
            preliminary_bid_price=None,
            gross_margin_amount=None,
            gross_margin_percent=None,
            logistics_reserve=None,
            risk_reserve=None,
            payment_delay_days=payment_delay_days,
            cash_gap_estimate=None,
            economics_status="insufficient_data",
            selected_supplier_name=None,
            assumptions={
                "target_margin_percent": target_margin_percent,
                "logistics_reserve_percent": logistics_reserve_percent,
                "risk_reserve_percent": risk_reserve_percent,
            },
            manual_checks=[ManualCheck(code="supplier_cost_missing", message="Нет распознанных ТКП для расчёта экономики.")],
            warnings=[],
            limitations=["Supplier quote totals are unavailable."],
        )

    priced_suppliers = [supplier for supplier in quote_comparison.suppliers if supplier.total_amount is not None]
    if not priced_suppliers:
        return EconomicsSummary(
            status="partial",
            analysis_mode=analysis_mode,
            currency=None,
            supplier_cost_min=None,
            supplier_cost_selected=None,
            expected_revenue=None,
            preliminary_bid_price=None,
            gross_margin_amount=None,
            gross_margin_percent=None,
            logistics_reserve=None,
            risk_reserve=None,
            payment_delay_days=payment_delay_days,
            cash_gap_estimate=None,
            economics_status="insufficient_data",
            selected_supplier_name=None,
            assumptions={
                "target_margin_percent": target_margin_percent,
                "logistics_reserve_percent": logistics_reserve_percent,
                "risk_reserve_percent": risk_reserve_percent,
            },
            manual_checks=[ManualCheck(code="supplier_price_missing", message="В ТКП не найдено достаточно цен для экономики.")],
            warnings=[],
            limitations=["Supplier price totals are missing or low-confidence."],
        )

    selected_supplier = min(priced_suppliers, key=lambda supplier: supplier.total_amount or float("inf"))
    supplier_cost_min = min(supplier.total_amount for supplier in priced_suppliers if supplier.total_amount is not None)
    supplier_cost_selected = selected_supplier.total_amount
    currency = selected_supplier.currency or next((supplier.currency for supplier in priced_suppliers if supplier.currency), None)
    logistics_reserve = round(supplier_cost_selected * (logistics_reserve_percent / 100), 2) if supplier_cost_selected else None
    risk_reserve = round(supplier_cost_selected * (risk_reserve_percent / 100), 2) if supplier_cost_selected else None
    base_cost = round((supplier_cost_selected or 0) + (logistics_reserve or 0) + (risk_reserve or 0), 2) if supplier_cost_selected is not None else None
    preliminary_bid_price = None
    if base_cost is not None and target_margin_percent < 100:
        preliminary_bid_price = round(base_cost / max(1 - (target_margin_percent / 100), 0.01), 2)

    cash_gap_estimate = round((supplier_cost_selected or 0) * min(payment_delay_days / 90, 1.5), 2) if supplier_cost_selected is not None else None
    warnings: list[ExtractionWarning] = []
    manual_checks: list[ManualCheck] = [
        ManualCheck(code="revenue_missing", message="Цена заказчика не найдена автоматически; expected revenue unavailable."),
    ]
    if quote_comparison.status != "completed":
        manual_checks.append(ManualCheck(code="quote_partial", message="Часть ТКП извлечена с ограничениями. Перед решением нужна ручная проверка."))
    if len({supplier.currency for supplier in priced_suppliers if supplier.currency}) > 1:
        warnings.append(ExtractionWarning(code="mixed_currency", message="Detected multiple currencies across supplier quotes."))
        manual_checks.append(ManualCheck(code="currency_normalization", message="Нормализовать валюты до финального расчёта экономики."))

    economics_status = "conditionally_viable" if preliminary_bid_price is not None else "insufficient_data"
    return EconomicsSummary(
        status="completed" if preliminary_bid_price is not None else "partial",
        analysis_mode=analysis_mode,
        currency=currency,
        supplier_cost_min=round(supplier_cost_min, 2) if supplier_cost_min is not None else None,
        supplier_cost_selected=round(supplier_cost_selected, 2) if supplier_cost_selected is not None else None,
        expected_revenue=None,
        preliminary_bid_price=preliminary_bid_price,
        gross_margin_amount=None,
        gross_margin_percent=target_margin_percent if preliminary_bid_price is not None else None,
        logistics_reserve=logistics_reserve,
        risk_reserve=risk_reserve,
        payment_delay_days=payment_delay_days,
        cash_gap_estimate=cash_gap_estimate,
        economics_status=economics_status,
        selected_supplier_name=selected_supplier.supplier_name,
        assumptions={
            "target_margin_percent": target_margin_percent,
            "logistics_reserve_percent": logistics_reserve_percent,
            "risk_reserve_percent": risk_reserve_percent,
        },
        manual_checks=manual_checks,
        warnings=warnings,
        limitations=[
            "Economics are demo-mode estimates built from uploaded quote totals and operator defaults.",
            "Expected revenue remains unavailable unless customer target price is provided separately.",
        ],
    )
