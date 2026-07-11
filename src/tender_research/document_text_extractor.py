from __future__ import annotations

import io
import os
import zipfile
from pathlib import Path
from xml.etree import ElementTree


EXTRACTED_STATUS = "extracted"
FAILED_STATUS = "failed"
UNSUPPORTED_STATUS = "unsupported"
EMPTY_STATUS = "empty"


def extract_text(local_path: str, max_chars: int = 2_000_000) -> tuple[str, str]:
    ext = Path(local_path).suffix.lower()
    try:
        with open(local_path, "rb") as f:
            content = f.read()
    except OSError as e:
        return FAILED_STATUS, f"File read error: {e}"
    if not content:
        return EMPTY_STATUS, ""
    result = _extract_by_ext(ext, content, max_chars)
    if result is None or not result.strip():
        if _is_unsupported_ext(ext):
            return UNSUPPORTED_STATUS, (result or "")
        return EMPTY_STATUS, (result or "")
    return EXTRACTED_STATUS, result[:max_chars]


def _extract_by_ext(ext: str, content: bytes, max_chars: int) -> str | None:
    if ext == ".txt":
        return _extract_txt(content)
    if ext == ".docx":
        return _extract_docx(content)
    if ext == ".pdf":
        return _extract_pdf(content, max_chars)
    if ext in (".xlsx", ".xls"):
        return _extract_xlsx(content)
    if ext in (".html", ".htm"):
        return _extract_html(content)
    if ext == ".xml":
        return _extract_xml(content)
    if ext == ".csv":
        return _extract_txt(content)
    if ext == ".json":
        return _extract_txt(content)
    return None


def _is_unsupported_ext(ext: str) -> bool:
    return ext not in (".txt", ".docx", ".pdf", ".xlsx", ".xls", ".html", ".htm", ".xml", ".csv", ".json")


def _extract_txt(content: bytes) -> str:
    for enc in ("utf-8", "cp1251", "koi8-r", "latin-1"):
        try:
            return content.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
    return content.decode("utf-8", errors="replace")


def _extract_docx(content: bytes) -> str:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as z:
            xml_content = z.read("word/document.xml")
        root = ElementTree.fromstring(xml_content)
        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        body = root.find("w:body", ns)
        if body is None:
            return ""

        blocks: list[str] = []
        for child in body:
            tag = child.tag.rsplit("}", 1)[-1]
            if tag == "p":
                paragraph_text = "".join(t.text or "" for t in child.iter("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t")).strip()
                if paragraph_text:
                    blocks.append(paragraph_text)
            elif tag == "tbl":
                for row in child.findall("w:tr", ns):
                    cells: list[str] = []
                    for cell in row.findall("w:tc", ns):
                        cell_text = "".join(t.text or "" for t in cell.iter("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t")).strip()
                        cells.append(cell_text)
                    normalized_cells = [cell for cell in cells if cell]
                    if normalized_cells:
                        blocks.append("\t".join(normalized_cells))
        return "\n".join(blocks)
    except Exception:
        return ""


def _extract_pdf(content: bytes, max_chars: int) -> str:
    try:
        import pypdf
    except ImportError:
        return ""
    try:
        reader = pypdf.PdfReader(io.BytesIO(content))
        texts = []
        for i, page in enumerate(reader.pages):
            if i >= 10:
                break
            t = page.extract_text() or ""
            texts.append(t)
        return " ".join(texts)[:max_chars]
    except Exception:
        return ""


def _extract_xlsx(content: bytes) -> str:
    try:
        import openpyxl
    except ImportError:
        return ""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        lines = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_lines = []
            for row in ws.iter_rows(values_only=True):
                vals = [str(v) if v is not None else "" for v in row]
                sheet_lines.append("\t".join(vals))
            if sheet_lines:
                lines.append(f"=== {sheet_name} ===")
                lines.extend(sheet_lines)
        wb.close()
        return "\n".join(lines)
    except Exception:
        return ""


def _extract_html(content: bytes) -> str:
    text = _extract_txt(content)
    import re
    clean = re.sub(r"<script[^>]*>.*?</script>", "", text, flags=re.DOTALL | re.IGNORECASE)
    clean = re.sub(r"<style[^>]*>.*?</style>", "", clean, flags=re.DOTALL | re.IGNORECASE)
    clean = re.sub(r"<[^>]+>", " ", clean)
    clean = re.sub(r"\s+", " ", clean).strip()
    return clean[:500_000]


def _extract_xml(content: bytes) -> str:
    try:
        text = content.decode("utf-8")
    except UnicodeDecodeError:
        text = content.decode("cp1251", errors="replace")
    import re
    clean = re.sub(r"<[^>]+>", " ", text)
    clean = re.sub(r"\s+", " ", clean).strip()
    return clean[:500_000]
